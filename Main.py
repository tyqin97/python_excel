import pandas as pd
from openpyxl import load_workbook, Workbook
import pyfiglet
import os
import uuid
from datetime import datetime
import numpy as np
from pprint import pprint

class DMT5():
    def __init__(self, path):
        super().__init__()
        
        self.path = path

        # Fix or Main Fix Data
        self.company = "GTI001"
        self.typecode = "M"
        self.costmethod = "F"
        self.nonstock = "TRUE"
        self.qtybearing = "TRUE"
        self.trackserialnum = "FALSE"
        self.buytoorder = "FALSE"
        self.plant = "MsgSys"
        self.oprseq = 10
        self.stdformat = "PH"
        self.prodstandard = 0
        self.qtyper = 1
        self.related_opt = 10
        self.eco_group = ""
        self.viewAsAsm = 1
        self.planAsAsm = 0

        # Initialize User Input
        self.main_drawing_number = ""
        self.main_eco_group_id = ""
        self.main_effective_date = ""
        self.main_total_module = ""
        self.main_total_fab_part = ""

        # Initialize Data Collection
        self.ext_li = []
        self.uom_dic = {}
        self.module_number = ""
        self.checkBOO = []

        # Initialize DMT5 Data
        self.part_master_data = []
        self.part_rev_data = []
        self.part_rev_attch_data = []
        self.part_boo_data = []
        self.part_bom_data = []

        # Get Current TImestamp
        self.curr_dt = datetime.now()
        self.timestamp = int(round(self.curr_dt.timestamp()))
        self.full_path = "./files/output/"+str(self.timestamp)

        # Create Folder with TimeStamp
        self.createFolder()
        # Print Out Logo
        self.logoPrint()
        # Load Excel Main Workbook
        self.loadWorkbook()
        # Get Info From Mainsheet
        self.getInputFromMainSheet()
        # Print Out Info
        self.showUserInput()
        # Get Data from all sheets
        self.getFrmAllSheets(uom_seq=True)
        # Get All Value
        self.getMainSheetVal()
        
    # Make directory for excel file output
    def createFolder(self):
        try:
            os.makedirs(self.full_path)
        except FileExistsError:
            pass

    def logoPrint(self):
        self.logo = pyfiglet.figlet_format("DMT5 Generator")
        print(self.logo)

    # Load Excel Workbook with path specified
    def loadWorkbook(self):
        print("Initializing System...")
        print("Reading Data...")

        self.wb_main = load_workbook(self.path, data_only=True)
        print('Reading Done!')

    # Get Value To Generate Report
    def getInputFromMainSheet(self):
        print("Getting Information from User")

        self.all_sheet_wb_main = []
        self.count_sheet_wb_main = 0

        for sheet in self.wb_main:
            if sheet.title.strip().upper() != "MAINSHEET":
                self.count_sheet_wb_main += 1
            self.all_sheet_wb_main.append(sheet)

        # Load MainSheet
        self.main_sheet = self.wb_main[self.all_sheet_wb_main[0].title]

    # Print Out Generated Report Before Run
    def showUserInput(self):
        self.main_drawing_number = self.main_sheet['B2'].value
        self.main_eco_group_id = self.main_sheet['B3'].value
        self.main_effective_date = str(self.main_sheet['B4'].value.strftime("%d-%b-%y"))
        self.main_total_module = self.main_sheet['B5'].value
        self.main_total_fab_part = self.main_sheet['B6'].value

        print("------------------MAIN INFO----------------------")
        print("Total Module: ", self.count_sheet_wb_main)
        print("Drawing Number: ", self.main_drawing_number)
        print("ECO Group ID: ", self.main_eco_group_id)
        print("Effective Date: ", self.main_effective_date)
        print("Total Module (Read Until): ", self.main_total_module)
        print("Total Fab Part: ", self.main_total_fab_part)

    # Split and Generate for PullAsAsm (BOM)
    def checkSplit(self, data):
        try:
            module_number_split = data.split("-")
            first_char = module_number_split[2][0]
            if first_char == "W":
                return "1"
            else:
                return "0"
        except:
            return "0"
    
    # Check list exist in another list or not
    def checkNestedList(self, li, item):
        if item in li:
            return True
        else:
            return False

    # Check the string splitable or not (BOM) 
    # Eliminate the abnormal data
    def checkSplitSucc(self, name):
        try:
            name.split("-")
        except:
            print(name)
            return False
        else:
            return True

    def getFrmAllSheets(self, bom=False, bom_main=True, uom_seq=False):
        check_li = []
        
        for sheet in self.all_sheet_wb_main:
            self.parts_li = []
            total_bom_li = []
            if sheet.title.strip().upper() == str(int(self.main_total_module) + 1):
                break
            
            if sheet.title.strip().upper() != "MAINSHEET" and bom == False:
                # Concat all sheets to map uom
                module_col = sheet['B2'].value
                module_uom = "SET"

                # Check if module already exist before
                if module_col not in check_li:
                    check_li.append(module_col)
                    self.uom_dic.update({module_col : module_uom})

                # Skip Top 2 Row and End Before Last Row in Column 3 Data
                ind = 0
                for data in sheet:
                    if ind != 2:
                        ind += 1
                        continue
                    if data[3].value == None:
                        break
                    
                    # Store All Data in Dictionary
                    part_drawing = data[3].value
                    part_erp = data[4].value
                    quantity = data[5].value
                    uom_code = data[6].value
                    revision = data[7].value

                    # Append all data into a list
                    self.parts_li.append([part_drawing, part_erp, quantity, uom_code, revision])
                self.ext_li.extend(self.parts_li)

            if sheet.title.strip().upper() != "MAINSHEET" and bom == True:
                # Run Main 
                if bom_main:
                    # print("Running Module Main...")
                    # Get the Module Number for every sheet before start
                    module_number = sheet["B2"].value

                    # Loop through sheets
                    skip_col = 0
                    mtl_seq = 10
                    for data in sheet:
                        if skip_col != 2:
                            skip_col += 1
                            continue
                        if data[3].value == None:
                            break

                        part_drawing = data[3].value
                        quantity = data[5].value
                        uom_code = data[6].value
                        rev_num = data[7].value

                        # Check the list for repitition return status False/ True
                        bom_status = self.checkNestedList(li=total_bom_li, item=[module_number, part_drawing, uom_code])

                        # If not exist before only do this
                        if not bom_status:
                            # Append to List for checking later
                            total_bom_li.append([module_number, part_drawing, uom_code])
                            # Get PullAsAsm Value
                            pullAsm = self.checkSplit(data=part_drawing)

                            if part_drawing != "xxx" and part_drawing != "xxxx":
                                # Append into BOM List
                                self.part_bom_data.append([self.company, self.plant, module_number, rev_num, mtl_seq, part_drawing, quantity, 
                                                        uom_code, self.related_opt, self.eco_group, self.viewAsAsm, pullAsm, self.planAsAsm])

                            # Increase MtlSeq for next loop
                            mtl_seq += 10
                        
                        else:
                            pass
                    

                    # print("Running Part Main...")
                    skip_col = 0
                    mtl_seq = 10
                    prev_main_part = ""
                    for data in sheet:
                        if skip_col != 2:
                            skip_col += 1
                            continue
                        if data[3].value == None:
                            break
                        # if data[4].value != "":
                        #     continue

                        main_part = data[3].value
                        sub_part = data[4].value
                        quantity = 1
                        uom_code = data[6].value
                        rev_num = data[7].value

                        # Check the list for repitition return status False / True
                        bom_status = self.checkNestedList(li=total_bom_li, item=[main_part, sub_part, uom_code])

                        # Check still in the same seq or not
                        if prev_main_part == main_part:
                            # If not exist before only do this
                            if not bom_status:
                                # Append to List for checking later
                                total_bom_li.append([main_part, sub_part, uom_code])
                                # Get PullAsm Value
                                pullAsm = self.checkSplit(data=sub_part)
                                # Check for abnormal string
                                splitSucc = self.checkSplitSucc(main_part)

                                if splitSucc == True:
                                    if sub_part != None and sub_part != "N/A" and sub_part != "xxx" and sub_part != "xxxx":
                                        # Append into BOM List
                                        self.part_bom_data.append([self.company, self.plant, main_part, rev_num, mtl_seq, sub_part, 1, uom_code,
                                                                self.related_opt, self.eco_group, self.viewAsAsm, pullAsm, self.planAsAsm])

                                # Increase MtlSeq for next loop
                                mtl_seq += 10

                        else:
                            # Replace new main part if not the same as previous
                            prev_main_part = main_part
                            # Reset the MtlSeq
                            mtl_seq = 10
                            # If not exist before only do this
                            if not bom_status:
                                # Append to List for checking later
                                total_bom_li.append([main_part, sub_part, uom_code])
                                # Get PullAsm Value
                                pullAsm = self.checkSplit(data=sub_part)
                                # Check for abnormal string
                                splitSucc = self.checkSplitSucc(main_part)

                                if splitSucc == True:
                                    if sub_part != None and sub_part != "N/A" and sub_part != "xxx" and sub_part != "xxxx":
                                        # Append into BOM List
                                        self.part_bom_data.append([self.company, self.plant, main_part, rev_num, mtl_seq, sub_part, 1, uom_code,
                                                                self.related_opt, self.eco_group, self.viewAsAsm, pullAsm, self.planAsAsm])

                                # Increase MtlSeq for next loop
                                mtl_seq += 10

        if uom_seq:
            df = self.convertLiToDf(self.ext_li, col_name=['Part Drawing', 'Part ERP Number', 'Quantity', 'UOMCode', 'Revision'])
            # Continue to concat all sheets with uom mapped
            for row in df.itertuples():
                main_col = row._1
                main_uom = row.UOMCode

                # Check Main either exist or not
                if main_col not in check_li:
                    check_li.append(main_col.replace("\n", ""))
                    self.uom_dic.update({main_col : main_uom})
                
                # Check Sec either exist or not
                else:
                    sec_col = row._2
                    if sec_col not in check_li and sec_col != "" and sec_col is not None:
                        check_li.append(sec_col.replace("\n", ""))
                        self.uom_dic.update({sec_col : main_uom})

    # Convert List of Data To Dataframe
    def convertLiToDf(self, li, col_name : list):
        df = pd.DataFrame(li, columns=col_name)
        return df

    # Convert Dataframe To Excel and Save To Directory
    def convertDfToExcel(self, df, sheet_name, file_name):
        df.to_excel(file_name, sheet_name, index=False)

    # Map the UOM using the dictionary generate in self.getFrmAllSheets (self.uom_dict)
    def mapUOM(self, col_name, ref_col):
        for col in col_name:
            self.part_master_df[col] = self.part_master_df[ref_col].map(self.uom_dic)

    # Split the Product Number by (-)
    def splitPartNum(self):
        self.part_num_split = self.part_num.split("-")
        self.get_char = self.part_num_split[2][0]
        self.setProdCode()
        
    # Generate ProdCode (Part Master)
    def setProdCode(self):
        if self.get_char == "U" or self.get_char == "W":
            self.ProdCode = "ASY"
            self.OPCode = "ASY"

        if self.get_char == "X":
            self.ProdCode = "FAB"
            self.OPCode = "FAB"

    # Generate RevShortDesc and PartAuditChangeDescription (Part Revision)
    def setDescription(self):
        if str(self.rev_numb) == "0":
            self.RevShortDesc = "INITIAL DESIGN"
            self.PartAuditChangeDescription = "INITIAL RELEASE"
        else:
            self.RevShortDesc = "UP REVISION"
            self.PartAuditChangeDescription = ""

    # Generate DrawDesc (Part Revision with Attachment)
    def chgDrawDesc(self):
        self.DrawDesc = self.main_drawing_number + "-R" + str(self.rev_numb)

    # Generate a list of sub part number (BOO, Part Revision, Part Revision with Attachment)
    def remBOOPart(self, pn):
        li = pn.split("-")
        if len(li) > 3:
            self.checkBOO.append(pn)

    # Loop Through the mainsheet
    def getMainSheetVal(self):
        ind = 0
        self.getBOMVal()
        for msheet in self.main_sheet:
            if ind != 2:
                ind += 1
                continue
            if msheet[3].value == None:
                break

            self.part_num = msheet[3].value
            self.part_desc = msheet[4].value
            self.rev_numb = msheet[5].value
            self.class_id = msheet[7].value
            self.file_dest = msheet[8].value

            self.mfrpart_numc = ""
            self.UOMClassID = "COUNT"
            self.IUM = ""
            self.SalesUM = ""
            self.PUM = ""
            self.ProdCode = ""

            self.PartBrand_c = "GREATECH"
            self.CommodityCode = ""
            self.NetWeight = ""
            self.NetWeightUOM = ""
            self.GrossWeight = ""
            self.GrossWeightUOM = ""
            self.RplPartNum_c = ""
            self.OPCode = ""
            self.DrawDesc = ""

            self.RevShortDesc = ""
            self.PartAuditChangeDescription = ""
            self.AltMethod = ""

            self.splitPartNum()
            self.chgDrawDesc()
            self.setDescription()
            self.remBOOPart(self.part_num)

            # Check whether sub part exist or not
            if self.part_num not in self.checkBOO:
                # Part BOO
                self.part_boo_data.append([self.company, self.plant, self.part_num, self.rev_numb, self.oprseq, 
                                        self.OPCode, self.main_eco_group_id, self.stdformat, self.prodstandard, self.qtyper])

                # Part Revision
                self.part_rev_data.append([self.company, self.part_num, self.rev_numb, self.RevShortDesc, self.main_effective_date,
                                           self.AltMethod, self.PartAuditChangeDescription, self.main_drawing_number])

                # Part Revision with Attachment
                self.part_rev_attch_data.append([self.company, self.part_num, self.rev_numb, self.file_dest, self.DrawDesc])

            # Part Master
            self.part_master_data.append([self.company, self.part_num, self.part_desc, 
                                      self.mfrpart_numc, self.typecode, self.UOMClassID, 
                                      self.IUM, self.SalesUM, self.PUM, self.ProdCode, self.class_id, 
                                      self.PartBrand_c, self.CommodityCode, self.NetWeight, self.NetWeightUOM, 
                                      self.GrossWeight, self.GrossWeightUOM, self.costmethod, self.nonstock, 
                                      self.qtybearing, self.trackserialnum, self.RplPartNum_c, self.buytoorder])           
                                    
    def getBOMVal(self):
        self.getFrmAllSheets(bom=True)


    def genPartMaster(self):
        self.part_master_df = self.convertLiToDf(self.part_master_data, col_name=['Company', 'PartNum', 'PartDescription', 'MfrPartNum_c', 
                                                                             'TypeCode', 'UOMClassID', 'IUM', 'SalesUM', 'PUM', 'ProdCode', 
                                                                             'ClassID', 'PartBrand_c', 'CommodityCode', 'NetWeight', 'NetWeightUOM', 
                                                                             'GrossWeight', 'GrossWeightUOM', 'CostMethod', 'NonStock', 'QtyBearing', 
                                                                             'TrackSerialNum', 'RplPartNum_c', 'BuyToOrder'])

        self.mapUOM(col_name=["IUM", "SalesUM", "PUM"], ref_col="PartNum")

        self.convertDfToExcel(self.part_master_df, sheet_name="Part", file_name=self.full_path + "/PartMaster.xlsx")

    def genPartRevision(self):
        self.part_rev_df = self.convertLiToDf(self.part_rev_data, col_name=['Company', 'PartNum', 'RevisionNum', 'RevShortDesc', 'EffectiveDate', 
                                                                            'AltMethod', 'PartAudit#ChangeDescription', 'DrawDesc'])

        self.convertDfToExcel(self.part_rev_df, sheet_name="Part Revision", file_name=self.full_path + "/PartRevision.xlsx")

    def genPartRevAttch(self):
        self.part_rev_attch_df = self.convertLiToDf(self.part_rev_attch_data, col_name=['Company', 'PartNum', 'RevisionNum', 'FileName', 'DrawNum'])

        self.convertDfToExcel(self.part_rev_attch_df, sheet_name="Part Revision With Attachment", file_name=self.full_path + "/PartRevisionWithAttachment.xlsx")

    def genBOO(self):
        self.part_boo_df = self.convertLiToDf(self.part_boo_data, col_name=['Company', 'Plant', 'PartNum', 'RevisionNum', 'OprSeq', 'OpCode', 
                                                                            'ECOGroupID', 'StdFormat', 'ProdStandard', 'QtyPer'])

        self.convertDfToExcel(self.part_boo_df, sheet_name="Bill Of Operations", file_name=self.full_path + "/BOO.xlsx")

    def genBOM(self):
        self.part_bom_df = self.convertLiToDf(self.part_bom_data, col_name=['Company', 'Plant', 'PartNum', 'RevisionNum', 'MtlSeq', 'MtlPartNum', 'QtyPer', 
                                                                            'UOMCode', 'RelatedOperation', 'ECOGroupID', 'ViewAsAsm', 'PullAsAsm', 'PlanAsAsm'])

        self.convertDfToExcel(self.part_bom_df, sheet_name="Bill Of Materials", file_name=self.full_path + "/BOM.xlsx")
        pass

    def selectGenFile(self, selectXL):
        if selectXL.upper() == "PARTMASTER":
            self.genPartMaster()
            pass
        if selectXL.upper() == "PARTREVISION":
            self.genPartRevision()
            pass
        if selectXL.upper() == "PARTREVISIONWITHATTACHMENT":
            self.genPartRevAttch()
            pass
        if selectXL.upper() == "BOO":
            self.genBOO()
            pass
        if selectXL.upper() == "BOM":
            self.genBOM()
            pass
        if selectXL.upper() == "ALL":
            self.genPartMaster()
            self.genPartRevision()
            self.genPartRevAttch()
            self.genBOO()
            self.genBOM()

DMT = DMT5(path='D:/DMT5/files/Mech Automated System(1001-U01000).xlsm')

DMT.selectGenFile("all")